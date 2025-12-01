from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Body
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from enum import Enum
import razorpay
from collections import defaultdict
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO as ExcelBytesIO

from decimal import Decimal, ROUND_HALF_UP


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-super-secret-jwt-key-change-in-production')
JWT_ALGORITHM = 'HS256'
ACCESS_TOKEN_EXPIRE_MINUTES = 60
REFRESH_TOKEN_EXPIRE_DAYS = 30

# Razorpay Configuration
RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID', 'rzp_test_demo')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET', 'demo_secret')
razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

# Default pricing settings for SuperAdmin
DEFAULT_PRICING_SETTINGS = {
    "base_monthly_inr": {
        "100": 1000,
        "200": 1800,
        "300": 2500,
        "400": 3200,
        "500": 3800,
        "1000": 7000
    },
    "tenure_discounts": {
        "monthly": 0,
        "quarterly": 0.05,
        "half_yearly": 0.10,
        "yearly": 0.15
    },
    "gst_rate": 0.18
}

# Create the main app
app = FastAPI(title="TrackMyMess API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# SuperAdmin Pricing Models
class PricingSettings(BaseModel):
    base_monthly_inr: Dict[str, float]
    tenure_discounts: Dict[str, float]
    gst_rate: float

class UpdatePricingSettings(BaseModel):
    base_monthly_inr: Optional[Dict[str, float]] = None
    tenure_discounts: Optional[Dict[str, float]] = None
    gst_rate: Optional[float] = None

# Tenant Billing Models
class QuoteRequest(BaseModel):
    capacity: int
    tenure: str

class CheckoutRequest(BaseModel):
    capacity: int
    tenure: str

class QuoteResponse(BaseModel):
    subtotal: float
    tax: float
    total: float
    months: int

# Enums
class UserRole(str, Enum):
    SUPER_ADMIN = "SUPER_ADMIN"
    ADMIN = "ADMIN"
    STAFF = "STAFF"

class SubscriptionPlan(str, Enum):
    FREE_TRIAL = "FREE_TRIAL"
    BASIC = "BASIC"
    STANDARD = "STANDARD"
    PREMIUM = "PREMIUM"
    ENTERPRISE = "ENTERPRISE"

class TenantStatus(str, Enum):
    ACTIVE = "ACTIVE"
    TRIAL = "TRIAL"
    EXPIRED = "EXPIRED"
    SUSPENDED = "SUSPENDED"

class Gender(str, Enum):
    MALE = "MALE"
    FEMALE = "FEMALE"
    OTHER = "OTHER"

class BillingType(str, Enum):
    PER_DAY = "PER_DAY"
    MONTHLY = "MONTHLY"

class PaymentMethod(str, Enum):
    RAZORPAY = "RAZORPAY"
    CASH = "CASH"
    UPI = "UPI"
    BANK_TRANSFER = "BANK_TRANSFER"

class PaymentStatus(str, Enum):
    PENDING = "PENDING"
    COMPLETED = "COMPLETED"
    FAILED = "FAILED"

class PauseRequestStatus(str, Enum):
    PENDING = "PENDING"
    APPROVED = "APPROVED"
    REJECTED = "REJECTED"

class StaffRole(str, Enum):
    MANAGER = "MANAGER"
    ATTENDANCE_OPERATOR = "ATTENDANCE_OPERATOR"
    ACCOUNTANT = "ACCOUNTANT"
    COOK = "COOK"
    HELPER = "HELPER"

class InvoiceStatus(str, Enum):
    PENDING = "PENDING"
    PARTIALLY_PAID = "PARTIALLY_PAID"
    PAID = "PAID"

# Plan Configuration
PLAN_CONFIG = {
    SubscriptionPlan.FREE_TRIAL: {
        "name": "Free Trial",
        "price": 0,
        "capacity": 50,
        "allow_downloads": False,
        "trial_days": 14
    },
    SubscriptionPlan.BASIC: {
        "name": "Basic",
        "price": 500,
        "capacity": 100,
        "allow_downloads": True
    },
    SubscriptionPlan.STANDARD: {
        "name": "Standard",
        "price": 1000,
        "capacity": 300,
        "allow_downloads": True
    },
    SubscriptionPlan.PREMIUM: {
        "name": "Premium",
        "price": 2000,
        "capacity": 999999,
        "allow_downloads": True
    },
    SubscriptionPlan.ENTERPRISE: {
        "name": "Enterprise",
        "price": 0,  # Custom pricing
        "capacity": 999999,
        "allow_downloads": True
    }
}


# Auth Models
class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: Dict[str, Any]

class RefreshTokenRequest(BaseModel):
    refresh_token: str

# User Models
class UserBase(BaseModel):
    username: str
    email: EmailStr
    full_name: str
    role: UserRole
    tenant_id: Optional[str] = None

class UserCreate(UserBase):
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    created_at: datetime
    is_active: bool = True

# Tenant/Client Models
class TenantCreate(BaseModel):
    mess_name: str
    email: EmailStr
    password: str
    mobile: str
    address: str
    
    @field_validator('mobile')
    def validate_mobile(cls, v):
        if not v.isdigit() or len(v) != 10:
            raise ValueError('Mobile must be 10 digits')
        return v
    
    @field_validator('password')
    def validate_password(cls, v):
        if len(v) < 8:
            raise ValueError('Password must be at least 8 characters')
        return v

class TenantUpdate(BaseModel):
    mess_name: Optional[str] = None
    owner_name: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None
    capacity: Optional[int] = None
    subscription_plan: Optional[SubscriptionPlan] = None
    status: Optional[TenantStatus] = None
    
    @field_validator('mobile')
    def validate_mobile(cls, v):
        if v is not None:
            if not v.isdigit() or len(v) != 10:
                raise ValueError('Mobile must be 10 digits')
        return v
    
    @field_validator('capacity')
    def validate_capacity(cls, v):
        if v is not None and v <= 0:
            raise ValueError('Capacity must be greater than 0')
        return v

class Tenant(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    mess_name: str
    owner_name: str
    email: EmailStr
    mobile: str
    address: str
    capacity: int
    subscription_plan: SubscriptionPlan
    status: TenantStatus
    subscription_start: datetime
    subscription_end: datetime
    customer_count: int = 0
    created_at: datetime

class PaymentUpdate(BaseModel):
    amount: Optional[float] = None
    payment_method: Optional[PaymentMethod] = None
    transaction_id: Optional[str] = None
    notes: Optional[str] = None
    payment_status: Optional[PaymentStatus] = None    

# Customer Models
class CustomerCreate(BaseModel):
    full_name: str
    gender: Gender
    mobile: str
    email: Optional[EmailStr] = None
    address: str
    joining_date: datetime
    meal_plan_id: str
    monthly_rate: float
    security_deposit: float = 0
    id_proof_type: Optional[str] = None
    id_proof_number: Optional[str] = None
    
    @field_validator('mobile')
    @classmethod
    def validate_mobile(cls, v):
        if not v.isdigit() or len(v) != 10:
            raise ValueError('Mobile must be 10 digits')
        return v

class CustomerUpdate(BaseModel):
    full_name: Optional[str] = None
    gender: Optional[Gender] = None
    mobile: Optional[str] = None
    email: Optional[EmailStr] = None
    address: Optional[str] = None
    meal_plan_id: Optional[str] = None
    monthly_rate: Optional[float] = None
    security_deposit: Optional[float] = None
    is_active: Optional[bool] = None
    
    @field_validator('mobile')
    @classmethod
    def validate_mobile(cls, v):
        if v is not None:
            if not v.isdigit() or len(v) != 10:
                raise ValueError('Mobile must be 10 digits')
        return v

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    full_name: str
    gender: Gender
    mobile: str
    email: Optional[EmailStr] = None
    address: str
    joining_date: datetime
    meal_plan_id: str
    meal_plan_name: Optional[str] = None
    monthly_rate: float
    security_deposit: float
    id_proof_type: Optional[str] = None
    id_proof_number: Optional[str] = None
    is_active: bool = True
    current_dues: float = 0
    created_at: datetime

# Meal Plan Models
class MealPlanCreate(BaseModel):
    name: str
    description: str
    meals_included: List[str]
    billing_type: BillingType
    rate: float
    is_active: bool = True

class MealPlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    name: str
    description: str
    meals_included: List[str]
    billing_type: BillingType
    rate: float
    is_active: bool
    is_default: bool = False
    customer_count: int = 0
    created_at: datetime

# Attendance Models
class AttendanceRecord(BaseModel):
    customer_id: str
    breakfast: bool = False
    lunch: bool = False
    dinner: bool = False

class AttendanceCreate(BaseModel):
    date: datetime
    records: List[AttendanceRecord]

class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    date: datetime
    customer_id: str
    customer_name: Optional[str] = None
    breakfast: bool
    lunch: bool
    dinner: bool
    created_at: datetime

# Pause Request Models
class PauseRequestCreate(BaseModel):
    customer_id: str
    start_date: datetime
    end_date: datetime
    reason: str

class PauseRequestUpdate(BaseModel):
    status: PauseRequestStatus
    admin_notes: Optional[str] = None

class PauseRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    customer_id: str
    customer_name: Optional[str] = None
    start_date: datetime
    end_date: datetime
    reason: str
    status: PauseRequestStatus
    admin_notes: Optional[str] = None
    created_at: datetime

# Payment Models
class PaymentCreate(BaseModel):
    customer_id: str
    amount: float
    payment_method: PaymentMethod
    transaction_id: Optional[str] = None
    notes: Optional[str] = None

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    customer_id: str
    customer_name: Optional[str] = None
    amount: float
    payment_method: PaymentMethod
    payment_status: PaymentStatus
    transaction_id: Optional[str] = None
    razorpay_order_id: Optional[str] = None
    razorpay_payment_id: Optional[str] = None
    notes: Optional[str] = None
    payment_date: datetime
    created_at: datetime

# Invoice Models
class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tenant_id: str
    customer_id: str
    customer_name: Optional[str] = None
    invoice_number: str
    month: int
    year: int
    total_days: int
    present_days: int
    pause_days: int
    meal_charges: float
    adjustments: float = 0
    total_amount: float
    paid_amount: float = 0
    due_amount: float
    status: Optional[InvoiceStatus] = InvoiceStatus.PENDING
    generated_at: datetime

class AdminDashboard(BaseModel):
    total_customers: int
    capacity: int
    today_attendance: Dict[str, int]
    monthly_revenue: Dict[str, float]
    meals_served_month: int
    today_pause_requests: int
    recent_payments: List[Payment]
    recent_pause_requests: List[PauseRequest]
    # New field expected by frontend: list of simple payment objects
    upcoming_payments: List[Dict[str, Any]] = []


# Utility Functions

# Pricing and billing helper functions
async def load_pricing_settings(db) -> dict:
    """Load pricing settings from DB or return defaults."""
    doc = await db.sa_settings.find_one({"key": "pricing"})
    if doc and "settings" in doc:
        return doc["settings"]
    return DEFAULT_PRICING_SETTINGS

async def get_pricing_settings_doc(db) -> dict:
    """Get full pricing settings document."""
    doc = await db.sa_settings.find_one({"key": "pricing"})
    if doc:
        return doc.get("settings", DEFAULT_PRICING_SETTINGS)
    return DEFAULT_PRICING_SETTINGS

async def upsert_pricing_settings(db, settings: dict):
    """Upsert pricing settings to DB."""
    await db.sa_settings.update_one(
        {"key": "pricing"},
        {"$set": {"key": "pricing", "settings": settings}},
        upsert=True
    )

def select_capacity_key(base_monthly_inr: dict, requested_capacity: int) -> str:
    """Select capacity key using nearest higher-or-equal rule."""
    if not base_monthly_inr:
        raise ValueError("No capacity tiers available")
    
    available_keys = sorted([int(k) for k in base_monthly_inr.keys()])
    
    # Find exact match
    if requested_capacity in available_keys:
        return str(requested_capacity)
    
    # Find nearest higher
    for key in available_keys:
        if key >= requested_capacity:
            return str(key)
    
    # Return maximum if none higher
    return str(max(available_keys))

def tenure_to_months(tenure: str) -> int:
    """Map tenure to number of months."""
    mapping = {
        "monthly": 1,
        "quarterly": 3,
        "half_yearly": 6,
        "yearly": 12
    }
    if tenure not in mapping:
        raise ValueError(f"Invalid tenure: {tenure}. Must be one of: monthly, quarterly, half_yearly, yearly")
    return mapping[tenure]

def compute_quote_for_capacity(settings: dict, capacity: int, tenure: str) -> dict:
    """Compute quote with subtotal, tax, total, and months."""
    base_monthly_inr = settings.get("base_monthly_inr", {})
    tenure_discounts = settings.get("tenure_discounts", {})
    gst_rate = settings.get("gst_rate", 0.18)
    
    if not base_monthly_inr:
        raise ValueError("No pricing tiers configured")
    
    try:
        months = tenure_to_months(tenure)
    except ValueError as e:
        raise ValueError(str(e))
    
    try:
        capacity_key = select_capacity_key(base_monthly_inr, capacity)
    except ValueError as e:
        raise ValueError(str(e))
    
    base_price_monthly = base_monthly_inr[capacity_key]
    discount = tenure_discounts.get(tenure, 0)
    
    # Use Decimal for accurate calculations
    base = Decimal(str(base_price_monthly))
    months_dec = Decimal(str(months))
    discount_dec = Decimal(str(discount))
    gst_dec = Decimal(str(gst_rate))
    
    subtotal = base * months_dec * (Decimal('1') - discount_dec)
    tax = subtotal * gst_dec
    total = subtotal + tax
    
    # Round to 2 decimals
    subtotal_rounded = float(subtotal.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    tax_rounded = float(tax.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    total_rounded = float(total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    
    return {
        "subtotal": subtotal_rounded,
        "tax": tax_rounded,
        "total": total_rounded,
        "months": months
    }

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire, "type": "access"})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expired")
    except jwt.JWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    payload = decode_token(token)
    
    if payload.get("type") != "access":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token type")
    
    user_id = payload.get("sub")
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    
    if not user_doc or not user_doc.get("is_active"):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found or inactive")

    # Check if token has been revoked
    revoked = await db.revoked_tokens.find_one({"token": token})
    if revoked:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token revoked")

    return user_doc

async def require_super_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Super admin access required")
    return current_user

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in [UserRole.SUPER_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    return current_user

async def get_tenant_info(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        return None
    
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No tenant associated")
    
    tenant_doc = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant_doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tenant not found")
    
    return tenant_doc


# Helper Functions for Payment-Invoice FIFO Sync
async def _apply_payment_to_invoices(
    db,
    tenant_id: str,
    customer_id: str,
    payment_amount: float,
    payment_id: str
) -> float:
    """Apply payment to customer's unpaid invoices using FIFO (oldest first)."""
    invoices = await db.invoices.find(
        {
            "tenant_id": tenant_id,
            "customer_id": customer_id,
            "due_amount": {"$gt": 0}
        },
        {"_id": 0}
    ).sort("generated_at", 1).to_list(1000)
    
    remaining_payment = payment_amount
    
    for invoice in invoices:
        if remaining_payment <= 0:
            break
        
        current_due = invoice.get("due_amount", 0)
        amount_to_apply = min(remaining_payment, current_due)
        
        if amount_to_apply > 0:
            new_paid_amount = invoice.get("paid_amount", 0) + amount_to_apply
            new_due_amount = max(0, invoice.get("total_amount", 0) - new_paid_amount)
            
            if new_due_amount == 0:
                new_status = InvoiceStatus.PAID
            elif new_paid_amount > 0:
                new_status = InvoiceStatus.PARTIALLY_PAID
            else:
                new_status = InvoiceStatus.PENDING
            
            await db.invoices.update_one(
                {"id": invoice["id"]},
                {"$set": {
                    "paid_amount": new_paid_amount,
                    "due_amount": new_due_amount,
                    "status": new_status
                }}
            )
            
            remaining_payment -= amount_to_apply
    
    total_dues = await _recalculate_customer_dues(db, customer_id, tenant_id)
    return total_dues


async def _recalculate_customer_dues(
    db,
    customer_id: str,
    tenant_id: str
) -> float:
    """Recalculate customer's total dues from all invoices."""
    invoices = await db.invoices.find(
        {"customer_id": customer_id, "tenant_id": tenant_id},
        {"_id": 0, "due_amount": 1}
    ).to_list(10000)
    
    total_dues = sum(inv.get("due_amount", 0) for inv in invoices)
    return total_dues


async def _rebuild_customer_invoices_from_payments(
    db,
    tenant_id: str,
    customer_id: str
) -> None:
    """Rebuild all invoice allocations for a customer from scratch."""
    # 1. Reset all invoices
    invoices = await db.invoices.find(
        {"customer_id": customer_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(10000)
    
    for inv in invoices:
        await db.invoices.update_one(
            {"id": inv["id"]},
            {"$set": {
                "paid_amount": 0,
                "due_amount": inv.get("total_amount", 0),
                "status": InvoiceStatus.PENDING
            }}
        )
    
    # 2. Get all COMPLETED payments in chronological order
    all_payments = await db.payments.find({
        "customer_id": customer_id,
        "tenant_id": tenant_id,
        "payment_status": PaymentStatus.COMPLETED
    }, {"_id": 0}).sort("payment_date", 1).to_list(10000)
    
    # 3. Reapply all payments in order
    for pmt in all_payments:
        await _apply_payment_to_invoices(
            db,
            tenant_id,
            customer_id,
            pmt["amount"],
            pmt["id"]
        )
    
    # 4. Update customer dues
    new_dues = await _recalculate_customer_dues(db, customer_id, tenant_id)
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"current_dues": new_dues}}
    )

# Initialize default data
async def initialize_default_data():
    # Check if super admin exists
    super_admin = await db.users.find_one({"role": UserRole.SUPER_ADMIN})
    if not super_admin:
        admin_id = str(uuid.uuid4())
        admin_doc = {
            "id": admin_id,
            "username": "superadmin",
            "email": "admin@trackmymess.com",
            "full_name": "Super Administrator",
            "password": hash_password("Admin@123"),
            "role": UserRole.SUPER_ADMIN,
            "tenant_id": None,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_doc)
        logger.info("Super admin created: superadmin / Admin@123")
    
    # Ensure revoked_tokens TTL index exists
    try:
        await db.revoked_tokens.create_index("expires_at", expireAfterSeconds=0)
    except Exception:
        pass
    
    # Clean up any legacy emergency_contact fields from existing customers
    try:
        result = await db.customers.update_many(
            {"emergency_contact": {"$exists": True}},
            {"$unset": {"emergency_contact": ""}}
        )
        if result.modified_count > 0:
            logger.info(f"Removed emergency_contact field from {result.modified_count} customer records")
    except Exception as e:
        logger.warning(f"Could not clean up emergency_contact fields: {e}")

# API Routes

# Auth Routes
@api_router.get("/sa/settings/pricing")
async def get_pricing_settings(current_user: dict = Depends(require_super_admin)):
    """Get pricing settings (SuperAdmin only)."""
    settings = await get_pricing_settings_doc(db)
    return settings

@api_router.patch("/sa/settings/pricing")
async def update_pricing_settings(
    updates: UpdatePricingSettings,
    current_user: dict = Depends(require_super_admin)
):
    """Update pricing settings (SuperAdmin only)."""
    current_settings = await get_pricing_settings_doc(db)
    
    update_data = updates.model_dump(exclude_none=True)
    
    if update_data:
        for key, value in update_data.items():
            current_settings[key] = value
        
        await upsert_pricing_settings(db, current_settings)
    
    return {"message": "Settings updated successfully"}

@api_router.get("/billing/plan")
async def get_current_plan(
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Get current tenant plan and status."""
    tenant_status = tenant.get("status", TenantStatus.ACTIVE)
    
    # Map TenantStatus to expected format
    status_map = {
        TenantStatus.ACTIVE: "active",
        TenantStatus.TRIAL: "trial",
        TenantStatus.EXPIRED: "expired",
        TenantStatus.SUSPENDED: "expired"
    }
    
    status_str = status_map.get(tenant_status, "active")
    
    # Get capacity from tenant or plan config
    capacity = tenant.get("capacity")
    if capacity is None:
        plan_config = PLAN_CONFIG.get(tenant.get("subscription_plan"), {})
        capacity = plan_config.get("capacity", 100)
    
    # Determine tenure based on subscription period (default to monthly)
    tenure = "monthly"
    try:
        sub_start = datetime.fromisoformat(tenant.get("subscription_start", datetime.now(timezone.utc).isoformat()))
        sub_end = datetime.fromisoformat(tenant.get("subscription_end", datetime.now(timezone.utc).isoformat()))
        days_diff = (sub_end - sub_start).days
        
        if 80 <= days_diff <= 100:
            tenure = "quarterly"
        elif 170 <= days_diff <= 190:
            tenure = "half_yearly"
        elif 350 <= days_diff <= 380:
            tenure = "yearly"
    except Exception:
        pass
    
    return {
        "current_plan": {
            "plan": {
                "capacity": capacity,
                "tenure": tenure
            },
            "status": status_str
        }
    }

@api_router.post("/billing/quote", response_model=QuoteResponse)
async def get_billing_quote(
    request: QuoteRequest,
    current_user: dict = Depends(require_admin)
):
    """Get pricing quote for capacity and tenure."""
    settings = await load_pricing_settings(db)
    
    try:
        quote = compute_quote_for_capacity(settings, request.capacity, request.tenure)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    return QuoteResponse(**quote)

@api_router.post("/billing/checkout")
async def create_checkout_session(
    request: CheckoutRequest,
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Create checkout session and Razorpay order."""
    settings = await load_pricing_settings(db)
    
    try:
        quote = compute_quote_for_capacity(settings, request.capacity, request.tenure)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    checkout_id = uuid.uuid4().hex
    now = datetime.now(timezone.utc)
    
    razorpay_order_id = None
    
    # Create Razorpay order if configured
    try:
        if razorpay_client:
            amount_paise = int(round(quote["total"] * 100))
            order = razorpay_client.order.create({
                "amount": amount_paise,
                "currency": "INR",
                "payment_capture": 1
            })
            razorpay_order_id = order.get("id")
    except Exception as e:
        logger.error(f"Razorpay order creation failed: {e}")
        # Continue without Razorpay order - still create checkout
    
    checkout_doc = {
        "id": checkout_id,
        "tenant_id": tenant["id"],
        "capacity": request.capacity,
        "tenure": request.tenure,
        "subtotal": quote["subtotal"],
        "tax": quote["tax"],
        "total": quote["total"],
        "months": quote["months"],
        "razorpay_order_id": razorpay_order_id,
        "created_at": now.isoformat(),
        "status": "pending"
    }
    
    await db.checkouts.insert_one(checkout_doc)
    
    return {
        "message": "Checkout session created. Redirecting to payment gateway...",
        "checkout_id": checkout_id,
        "razorpay_order_id": razorpay_order_id
    }
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(request: LoginRequest):
    user_doc = await db.users.find_one({"email": request.email}, {"_id": 0})

    if not user_doc or not verify_password(request.password, user_doc["password"]):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
    
    if not user_doc.get("is_active"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Account is inactive")
    
    token_data = {"sub": user_doc["id"], "role": user_doc["role"], "tenant_id": user_doc.get("tenant_id")}
    access_token = create_access_token(token_data)
    refresh_token = create_refresh_token(token_data)
    
    user_data = {k: v for k, v in user_doc.items() if k != "password"}
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user=user_data
    )


@api_router.post("/auth/logout")
async def logout(refresh_token: Optional[str] = Body(None), access_token: Optional[str] = Body(None)):
    """Revoke provided tokens (access and/or refresh)."""
    now = datetime.now(timezone.utc)
    revoked_count = 0

    for token in [refresh_token, access_token]:
        if not token:
            continue
        try:
            payload = decode_token(token)
        except HTTPException:
            continue

        exp = payload.get("exp")
        try:
            if isinstance(exp, (int, float)):
                expires_at = datetime.fromtimestamp(int(exp), timezone.utc)
            else:
                expires_at = datetime.fromisoformat(exp)
        except Exception:
            expires_at = now + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)

        doc = {
            "token": token,
            "type": payload.get("type"),
            "user_id": payload.get("sub"),
            "revoked_at": now.isoformat(),
            "expires_at": expires_at.isoformat()
        }
        try:
            await db.revoked_tokens.insert_one(doc)
            revoked_count += 1
        except Exception:
            continue

    return {"message": f"Logged out ({revoked_count} token(s) revoked)"}

@api_router.post("/auth/refresh", response_model=TokenResponse)
async def refresh_token(request: RefreshTokenRequest):
    payload = decode_token(request.refresh_token)
    
    if payload.get("type") != "refresh":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token type")

    revoked = await db.revoked_tokens.find_one({"token": request.refresh_token})
    if revoked:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token revoked")
    
    user_id = payload.get("sub")
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    
    if not user_doc or not user_doc.get("is_active"):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found or inactive")
    
    token_data = {"sub": user_doc["id"], "role": user_doc["role"], "tenant_id": user_doc.get("tenant_id")}
    access_token = create_access_token(token_data)
    new_refresh_token = create_refresh_token(token_data)
    
    user_data = {k: v for k, v in user_doc.items() if k != "password"}
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=new_refresh_token,
        user=user_data
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return {k: v for k, v in current_user.items() if k != "password"}

# Super Admin Routes

@api_router.get("/super-admin/dashboard", response_model=SuperAdminDashboard)
async def get_super_admin_dashboard(current_user: dict = Depends(require_super_admin)):
    tenants = await db.tenants.find({}, {"_id": 0}).to_list(10000)
    
    total_clients = len(tenants)
    active_clients = len([t for t in tenants if t["status"] == TenantStatus.ACTIVE])
    trial_clients = len([t for t in tenants if t["status"] == TenantStatus.TRIAL])
    expired_clients = len([t for t in tenants if t["status"] == TenantStatus.EXPIRED])
    suspended_clients = len([t for t in tenants if t["status"] == TenantStatus.SUSPENDED])
    
    total_customers = sum(t.get("customer_count", 0) for t in tenants)
    
    mrr = sum(
        PLAN_CONFIG[t["subscription_plan"]]["price"]
        for t in tenants
        if t["status"] == TenantStatus.ACTIVE
    )
    
    now = datetime.now(timezone.utc)
    seven_days = now + timedelta(days=7)
    expiring_soon = len([
        t for t in tenants
        if t["status"] == TenantStatus.ACTIVE and
        datetime.fromisoformat(t["subscription_end"]) <= seven_days
    ])
    
    return SuperAdminDashboard(
        total_clients=total_clients,
        active_clients=active_clients,
        trial_clients=trial_clients,
        expired_clients=expired_clients,
        suspended_clients=suspended_clients,
        total_customers=total_customers,
        mrr=mrr,
        expiring_soon=expiring_soon
    )

@api_router.post("/super-admin/tenants", response_model=Tenant)
async def create_tenant(tenant_data: TenantCreate, current_user: dict = Depends(require_super_admin)):
    existing_tenant = await db.tenants.find_one({"email": tenant_data.email})
    existing_user = await db.users.find_one({"email": tenant_data.email})
    if existing_tenant or existing_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already associated with a mess")

    tenant_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    default_plan = SubscriptionPlan.BASIC
    default_capacity = PLAN_CONFIG.get(default_plan, {}).get("capacity", 100)

    owner_name = tenant_data.email.split('@')[0].replace('.', ' ').title()

    tenant_doc = {
        "id": tenant_id,
        "mess_name": tenant_data.mess_name,
        "owner_name": owner_name,
        "email": tenant_data.email,
        "mobile": tenant_data.mobile,
        "address": tenant_data.address,
        "capacity": default_capacity,
        "subscription_plan": default_plan,
        "status": TenantStatus.ACTIVE,
        "subscription_start": now.isoformat(),
        "subscription_end": (now + timedelta(days=30)).isoformat(),
        "customer_count": 0,
        "created_at": now.isoformat()
    }
    await db.tenants.insert_one(tenant_doc)

    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "username": tenant_data.email,
        "email": tenant_data.email,
        "full_name": owner_name,
        "password": hash_password(tenant_data.password),
        "role": UserRole.ADMIN,
        "tenant_id": tenant_id,
        "is_active": True,
        "created_at": now.isoformat()
    }
    await db.users.insert_one(user_doc)
    
    default_plans = [
        {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "name": "Full Board (3 meals)",
            "description": "Breakfast, Lunch, and Dinner",
            "meals_included": ["Breakfast", "Lunch", "Dinner"],
            "billing_type": BillingType.MONTHLY,
            "rate": 3000,
            "is_active": True,
            "is_default": True,
            "customer_count": 0,
            "created_at": now.isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "name": "Lunch + Dinner",
            "description": "Two meals per day",
            "meals_included": ["Lunch", "Dinner"],
            "billing_type": BillingType.MONTHLY,
            "rate": 2100,
            "is_active": True,
            "is_default": True,
            "customer_count": 0,
            "created_at": now.isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "name": "Only Lunch",
            "description": "Single meal - Lunch only",
            "meals_included": ["Lunch"],
            "billing_type": BillingType.MONTHLY,
            "rate": 1200,
            "is_active": True,
            "is_default": True,
            "customer_count": 0,
            "created_at": now.isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "name": "Only Dinner",
            "description": "Single meal - Dinner only",
            "meals_included": ["Dinner"],
            "billing_type": BillingType.MONTHLY,
            "rate": 1350,
            "is_active": True,
            "is_default": True,
            "customer_count": 0,
            "created_at": now.isoformat()
        }
    ]
    await db.meal_plans.insert_many(default_plans)
    
    return Tenant(**tenant_doc)

@api_router.get("/super-admin/tenants", response_model=List[Tenant])
async def get_tenants(current_user: dict = Depends(require_super_admin)):
    tenants = await db.tenants.find({}, {"_id": 0}).to_list(10000)
    return [Tenant(**t) for t in tenants]

@api_router.get("/super-admin/tenants/{tenant_id}", response_model=Tenant)
async def get_tenant(tenant_id: str, current_user: dict = Depends(require_super_admin)):
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tenant not found")
    return Tenant(**tenant)

@api_router.put("/super-admin/tenants/{tenant_id}", response_model=Tenant)
async def update_tenant(tenant_id: str, updates: TenantUpdate, current_user: dict = Depends(require_super_admin)):
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tenant not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    if update_data:
        await db.tenants.update_one({"id": tenant_id}, {"$set": update_data})
    
    updated_tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    return Tenant(**updated_tenant)

# Admin/Client Routes
@api_router.get("/admin/dashboard", response_model=AdminDashboard)
async def get_admin_dashboard(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    tenant_id = tenant["id"]

    # --- BASIC METRICS ---
    customers = await db.customers.find({"tenant_id": tenant_id, "is_active": True}, {"_id": 0}).to_list(10000)
    total_customers = len(customers)

    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    attendance_records = await db.attendance.find({
        "tenant_id": tenant_id,
        "date": {"$gte": today.isoformat(), "$lt": (today + timedelta(days=1)).isoformat()}
    }, {"_id": 0}).to_list(10000)
    present = len(set(a["customer_id"] for a in attendance_records if a.get("breakfast") or a.get("lunch") or a.get("dinner")))
    absent = total_customers - present

    current_month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    payments = await db.payments.find({
        "tenant_id": tenant_id,
        "payment_date": {"$gte": current_month_start.isoformat()},
        "payment_status": PaymentStatus.COMPLETED
    }, {"_id": 0}).to_list(10000)
    collected = sum(p["amount"] for p in payments)

    invoices = await db.invoices.find({
        "tenant_id": tenant_id,
        "month": current_month_start.month,
        "year": current_month_start.year
    }, {"_id": 0}).to_list(10000)
    total_due = sum(inv.get("due_amount", 0) for inv in invoices)

    month_attendance = await db.attendance.find({
        "tenant_id": tenant_id,
        "date": {"$gte": current_month_start.isoformat()}
    }, {"_id": 0}).to_list(10000)
    meals_served = sum(
        (1 if a.get("breakfast") else 0) +
        (1 if a.get("lunch") else 0) +
        (1 if a.get("dinner") else 0)
        for a in month_attendance
    )

    pause_requests_today = await db.pause_requests.find({
        "tenant_id": tenant_id,
        "status": PauseRequestStatus.PENDING,
        "created_at": {"$gte": today.isoformat()}
    }, {"_id": 0}).to_list(10000)

    # --- RECENT PAYMENTS (completed only) ---
    recent_payments_docs = await db.payments.find(
        {"tenant_id": tenant_id, "payment_status": PaymentStatus.COMPLETED},
        {"_id": 0}
    ).sort("created_at", -1).limit(5).to_list(5)
    recent_payments = [Payment(**p) for p in recent_payments_docs]

    # --- RECENT PAUSE REQUESTS ---
    recent_pause_docs = await db.pause_requests.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(5).to_list(5)
    recent_pause_requests = [PauseRequest(**p) for p in recent_pause_docs]

    # --- UPCOMING & PENDING PAYMENTS ---
    invoice_docs = await db.invoices.find(
        {"tenant_id": tenant_id, "due_amount": {"$gt": 0}},
        {"_id": 0}
    ).to_list(500)

    now = datetime.now(timezone.utc)
    payment_list = []
    for inv in invoice_docs:
        due_date_str = inv.get("due_date") or inv.get("generated_at")
        if not due_date_str:
            continue
        try:
            due_date = datetime.fromisoformat(due_date_str.replace("Z", "+00:00"))
        except Exception:
            continue

        payment_list.append({
            "id": inv.get("id"),
            "customer_name": inv.get("customer_name", ""),
            "due_date": due_date.isoformat(),
            "amount": inv.get("due_amount", 0),
            "is_pending": due_date < now
        })

    # Sort by due_date (oldest first)
    payment_list.sort(key=lambda x: x["due_date"])

    # Split into pending & upcoming, pick 5 each
    pending = [p for p in payment_list if p["is_pending"]][:5]
    upcoming = [p for p in payment_list if not p["is_pending"]][:5]

    upcoming_payments = pending + upcoming

    # --- RETURN DASHBOARD DATA ---
    return AdminDashboard(
        total_customers=total_customers,
        capacity=tenant["capacity"],
        today_attendance={"present": present, "absent": absent},
        monthly_revenue={"collected": collected, "pending": total_due},
        meals_served_month=meals_served,
        today_pause_requests=len(pause_requests_today),
        recent_payments=recent_payments,
        recent_pause_requests=recent_pause_requests,
        upcoming_payments=upcoming_payments
    )



# Customer Routes
@api_router.post("/admin/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    tenant_id = tenant["id"]
    
    capacity = tenant.get("capacity")
    if capacity is None:
        plan_config = PLAN_CONFIG.get(tenant.get("subscription_plan"), {})
        capacity = plan_config.get("capacity", 0)

    updated = await db.tenants.update_one(
        {"id": tenant_id, "customer_count": {"$lt": capacity}},
        {"$inc": {"customer_count": 1}}
    )

    if updated.modified_count == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot add student: Mess capacity reached."
        )
    
    existing = await db.customers.find_one({"tenant_id": tenant_id, "mobile": customer_data.mobile})
    if existing:
        # Rollback count increment
        await db.tenants.update_one({"id": tenant_id}, {"$inc": {"customer_count": -1}})
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Mobile number already exists")
    
    meal_plan = await db.meal_plans.find_one({"id": customer_data.meal_plan_id, "tenant_id": tenant_id})
    if not meal_plan:
        # Rollback count increment
        await db.tenants.update_one({"id": tenant_id}, {"$inc": {"customer_count": -1}})
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Meal plan not found")
    
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    customer_doc = {
        "id": customer_id,
        "tenant_id": tenant_id,
        **customer_data.model_dump(),
        "joining_date": customer_data.joining_date.isoformat(),
        "is_active": True,
        "current_dues": 0,
        "created_at": now.isoformat()
    }

    try:
        await db.customers.insert_one(customer_doc)
    except Exception as e:
        await db.tenants.update_one({"id": tenant_id}, {"$inc": {"customer_count": -1}})
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))

    try:
        await db.meal_plans.update_one({"id": customer_data.meal_plan_id}, {"$inc": {"customer_count": 1}})
    except Exception:
        pass

    customer_doc["meal_plan_name"] = meal_plan["name"]
    return Customer(**customer_doc)

@api_router.get("/admin/customers", response_model=List[Customer])
async def get_customers(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customers = await db.customers.find({"tenant_id": tenant["id"]}, {"_id": 0}).to_list(10000)
    
    for customer in customers:
        meal_plan = await db.meal_plans.find_one({"id": customer["meal_plan_id"]}, {"_id": 0})
        if meal_plan:
            customer["meal_plan_name"] = meal_plan["name"]
    
    return [Customer(**c) for c in customers]

@api_router.get("/admin/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tenant["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    meal_plan = await db.meal_plans.find_one({"id": customer["meal_plan_id"]}, {"_id": 0})
    if meal_plan:
        customer["meal_plan_name"] = meal_plan["name"]
    
    return Customer(**customer)

@api_router.put("/admin/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, updates: CustomerUpdate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tenant["id"]})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    if "meal_plan_id" in update_data and update_data["meal_plan_id"] != customer["meal_plan_id"]:
        await db.meal_plans.update_one({"id": customer["meal_plan_id"]}, {"$inc": {"customer_count": -1}})
        await db.meal_plans.update_one({"id": update_data["meal_plan_id"]}, {"$inc": {"customer_count": 1}})
    
    if update_data:
        await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    
    updated_customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    meal_plan = await db.meal_plans.find_one({"id": updated_customer["meal_plan_id"]}, {"_id": 0})
    if meal_plan:
        updated_customer["meal_plan_name"] = meal_plan["name"]
    
    return Customer(**updated_customer)

@api_router.delete("/admin/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    """Delete a customer and decrement tenant customer_count."""
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tenant["id"]})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    # Delete customer
    await db.customers.delete_one({"id": customer_id})
    
    # Decrement tenant customer count
    await db.tenants.update_one({"id": tenant["id"]}, {"$inc": {"customer_count": -1}})
    
    # Decrement meal plan customer count
    try:
        await db.meal_plans.update_one({"id": customer["meal_plan_id"]}, {"$inc": {"customer_count": -1}})
    except Exception:
        pass
    
    return {"message": "Customer deleted successfully"}


@api_router.get("/admin/payments/upcoming")
async def get_upcoming_payments(limit: int = 20, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    tenant_id = tenant["id"]
    invoice_docs = await db.invoices.find(
        {"tenant_id": tenant_id, "due_amount": {"$gt": 0}},
        {"_id": 0}
    ).sort("generated_at", 1).limit(limit).to_list(limit)

    result = []
    for inv in invoice_docs:
        due_date = inv.get("due_date") or inv.get("generated_at") or datetime.now(timezone.utc).isoformat()
        result.append({
            "id": inv.get("id"),
            "customer_name": inv.get("customer_name", ""),
            "due_date": due_date,
            "amount": inv.get("due_amount", 0)
        })
    return result


@api_router.get("/admin/payments/pending")
async def get_pending_payments(limit: int = 20, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    tenant_id = tenant["id"]
    now_iso = datetime.now(timezone.utc).isoformat()

    # Consider a payment pending if the computed due_date is in the past.
    # We look for invoices with due_amount > 0 and whose due_date/generated_at <= now.
    invoice_docs = await db.invoices.find({
        "tenant_id": tenant_id,
        "due_amount": {"$gt": 0},
        "$expr": {
            # we can't compare iso strings reliably in Mongo without storing as date,
            # so we will fetch and filter in Python for safety (below).
            # Keep initial filter minimal (tenant + due_amount).
        }
    }, {"_id": 0}).to_list(1000)

    result = []
    for inv in invoice_docs:
        due_date_str = inv.get("due_date") or inv.get("generated_at")
        if not due_date_str:
            continue
        try:
            due_dt = datetime.fromisoformat(due_date_str)
        except Exception:
            # if parsing fails, skip this invoice
            continue
        if due_dt <= datetime.now(timezone.utc):
            result.append({
                "id": inv.get("id"),
                "customer_name": inv.get("customer_name", ""),
                "due_date": due_date_str,
                "amount": inv.get("due_amount", 0)
            })
        if len(result) >= limit:
            break

    # Sort pending by due_date ascending (oldest due first)
    result.sort(key=lambda x: x["due_date"])
    return result


# Meal Plan Routes
@api_router.get("/admin/meal-plans", response_model=List[MealPlan])
async def get_meal_plans(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    plans = await db.meal_plans.find({"tenant_id": tenant["id"]}, {"_id": 0}).to_list(10000)
    return [MealPlan(**p) for p in plans]

@api_router.post("/admin/meal-plans", response_model=MealPlan)
async def create_meal_plan(plan_data: MealPlanCreate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    plan_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    plan_doc = {
        "id": plan_id,
        "tenant_id": tenant["id"],
        **plan_data.model_dump(),
        "is_default": False,
        "customer_count": 0,
        "created_at": now.isoformat()
    }
    
    await db.meal_plans.insert_one(plan_doc)
    return MealPlan(**plan_doc)

@api_router.put("/admin/meal-plans/{plan_id}", response_model=MealPlan)
async def update_meal_plan(plan_id: str, updates: MealPlanCreate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    plan = await db.meal_plans.find_one({"id": plan_id, "tenant_id": tenant["id"]})
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Meal plan not found")
    
    update_data = updates.model_dump()
    await db.meal_plans.update_one({"id": plan_id}, {"$set": update_data})
    
    updated_plan = await db.meal_plans.find_one({"id": plan_id}, {"_id": 0})
    return MealPlan(**updated_plan)

@api_router.delete("/admin/meal-plans/{plan_id}")
async def delete_meal_plan(plan_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    plan = await db.meal_plans.find_one({"id": plan_id, "tenant_id": tenant["id"]})
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Meal plan not found")
    
    if plan.get("customer_count", 0) > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{plan['customer_count']} customers on this plan. Reassign first."
        )
    
    await db.meal_plans.delete_one({"id": plan_id})
    return {"message": "Meal plan deleted successfully"}

# Attendance Routes
@api_router.post("/admin/attendance")
async def mark_attendance(attendance_data: AttendanceCreate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    tenant_id = tenant["id"]
    
    date_str = attendance_data.date.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    await db.attendance.delete_many({"tenant_id": tenant_id, "date": date_str})
    
    now = datetime.now(timezone.utc)
    records = []
    
    for record in attendance_data.records:
        customer = await db.customers.find_one({"id": record.customer_id, "tenant_id": tenant_id})
        if customer:
            doc = {
                "id": str(uuid.uuid4()),
                "tenant_id": tenant_id,
                "date": date_str,
                "customer_id": record.customer_id,
                "customer_name": customer["full_name"],
                "breakfast": record.breakfast,
                "lunch": record.lunch,
                "dinner": record.dinner,
                "created_at": now.isoformat()
            }
            records.append(doc)
    
    if records:
        await db.attendance.insert_many(records)
    
    return {"message": f"Attendance marked for {len(records)} customers"}

@api_router.get("/admin/attendance", response_model=List[Attendance])
async def get_attendance(date: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    date_obj = datetime.fromisoformat(date).replace(hour=0, minute=0, second=0, microsecond=0)
    
    records = await db.attendance.find({
        "tenant_id": tenant["id"],
        "date": date_obj.isoformat()
    }, {"_id": 0}).to_list(10000)
    
    return [Attendance(**r) for r in records]

# Pause Request Routes
@api_router.post("/admin/pause-requests", response_model=PauseRequest)
async def create_pause_request(pause_data: PauseRequestCreate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customer = await db.customers.find_one({"id": pause_data.customer_id, "tenant_id": tenant["id"]})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    pause_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    pause_doc = {
        "id": pause_id,
        "tenant_id": tenant["id"],
        "customer_id": pause_data.customer_id,
        "customer_name": customer["full_name"],
        "start_date": pause_data.start_date.isoformat(),
        "end_date": pause_data.end_date.isoformat(),
        "reason": pause_data.reason,
        "status": PauseRequestStatus.PENDING,
        "admin_notes": None,
        "created_at": now.isoformat()
    }
    
    await db.pause_requests.insert_one(pause_doc)
    return PauseRequest(**pause_doc)

@api_router.get("/admin/pause-requests", response_model=List[PauseRequest])
async def get_pause_requests(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    requests = await db.pause_requests.find({"tenant_id": tenant["id"]}, {"_id": 0}).to_list(10000)
    return [PauseRequest(**r) for r in requests]

@api_router.put("/admin/pause-requests/{pause_id}", response_model=PauseRequest)
async def update_pause_request(pause_id: str, updates: PauseRequestUpdate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    pause_req = await db.pause_requests.find_one({"id": pause_id, "tenant_id": tenant["id"]})
    if not pause_req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pause request not found")
    
    update_data = updates.model_dump()
    await db.pause_requests.update_one({"id": pause_id}, {"$set": update_data})
    
    updated = await db.pause_requests.find_one({"id": pause_id}, {"_id": 0})
    return PauseRequest(**updated)

# Payment Routes
@api_router.post("/admin/payments/razorpay/create-order")
async def create_razorpay_order(amount: float = Body(...), customer_id: str = Body(...), current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    try:
        order = razorpay_client.order.create({
            "amount": int(amount * 100),
            "currency": "INR",
            "payment_capture": 1
        })
        
        return {
            "order_id": order["id"],
            "amount": amount,
            "currency": "INR",
            "key_id": RAZORPAY_KEY_ID
        }
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

@api_router.post("/admin/payments", response_model=Payment)
async def create_payment(payment_data: PaymentCreate, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customer = await db.customers.find_one({"id": payment_data.customer_id, "tenant_id": tenant["id"]})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    payment_doc = {
        "id": payment_id,
        "tenant_id": tenant["id"],
        "customer_id": payment_data.customer_id,
        "customer_name": customer["full_name"],
        "amount": payment_data.amount,
        "payment_method": payment_data.payment_method,
        "payment_status": PaymentStatus.COMPLETED,
        "transaction_id": payment_data.transaction_id,
        "notes": payment_data.notes,
        "payment_date": now.isoformat(),
        "created_at": now.isoformat()
    }
    
    await db.payments.insert_one(payment_doc)
    
    current_dues = customer.get("current_dues", 0)
    new_dues = max(0, current_dues - payment_data.amount)
    await db.customers.update_one({"id": payment_data.customer_id}, {"$set": {"current_dues": new_dues}})
    
    return Payment(**payment_doc)
@api_router.get("/admin/payments", response_model=List[Payment])
async def get_payments(
    searchQuery: Optional[str] = None,
    paymentMethod: Optional[str] = None,
    paymentStatus: Optional[str] = None,
    dateFrom: Optional[str] = None,
    dateTo: Optional[str] = None,
    minAmount: Optional[float] = None,
    maxAmount: Optional[float] = None,
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Get payments with advanced filtering support."""
    query = {"tenant_id": tenant["id"]}
    
    # Apply filters
    if searchQuery:
        search_lower = searchQuery.lower()
        customers = await db.customers.find(
            {"tenant_id": tenant["id"]},
            {"_id": 0, "id": 1, "full_name": 1}
        ).to_list(10000)
        
        matching_customer_ids = [
            c["id"] for c in customers 
            if search_lower in c["full_name"].lower()
        ]
        
        query["$or"] = [
            {"customer_id": {"$in": matching_customer_ids}},
            {"transaction_id": {"$regex": searchQuery, "$options": "i"}},
            {"razorpay_payment_id": {"$regex": searchQuery, "$options": "i"}},
            {"notes": {"$regex": searchQuery, "$options": "i"}}
        ]
    
    if paymentMethod and paymentMethod != "ALL":
        query["payment_method"] = paymentMethod
    
    if paymentStatus and paymentStatus != "ALL":
        query["payment_status"] = paymentStatus
    
    if dateFrom:
        date_from_obj = datetime.fromisoformat(dateFrom).replace(hour=0, minute=0, second=0, microsecond=0)
        query["payment_date"] = {"$gte": date_from_obj.isoformat()}
    
    if dateTo:
        date_to_obj = datetime.fromisoformat(dateTo).replace(hour=23, minute=59, second=59, microsecond=999999)
        if "payment_date" in query:
            query["payment_date"]["$lte"] = date_to_obj.isoformat()
        else:
            query["payment_date"] = {"$lte": date_to_obj.isoformat()}
    
    if minAmount is not None:
        query["amount"] = {"$gte": minAmount}
    
    if maxAmount is not None:
        if "amount" in query:
            query["amount"]["$lte"] = maxAmount
        else:
            query["amount"] = {"$lte": maxAmount}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
    return [Payment(**p) for p in payments]

@api_router.put("/admin/payments/{payment_id}", response_model=Payment)
async def update_payment(
    payment_id: str,
    updates: PaymentUpdate,
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Update a payment record with proper dues recalculation."""
    payment = await db.payments.find_one({"id": payment_id, "tenant_id": tenant["id"]})
    if not payment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Payment not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid fields to update")
    
    # Handle amount change - recalculate customer dues
    if "amount" in update_data and payment.get("customer_id"):
        customer = await db.customers.find_one({"id": payment["customer_id"]})
        if customer:
            old_amount = payment["amount"]
            new_amount = update_data["amount"]
            difference = new_amount - old_amount
            
            # Adjust dues: if payment increased, reduce dues; if decreased, increase dues
            new_dues = max(0, customer.get("current_dues", 0) - difference)
            await db.customers.update_one(
                {"id": payment["customer_id"]},
                {"$set": {"current_dues": new_dues}}
            )
    
    await db.payments.update_one({"id": payment_id}, {"$set": update_data})
    updated_payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    
    # Rebuild allocations if amount/customer changed for COMPLETED payment
    old_customer_id = payment.get("customer_id")
    old_status = payment.get("payment_status")
    needs_rebuild = ("amount" in update_data or "customer_id" in update_data) and old_status == PaymentStatus.COMPLETED
    if needs_rebuild and old_customer_id:
        await _rebuild_customer_invoices_from_payments(db, tenant["id"], old_customer_id)
    
    return Payment(**updated_payment)

@api_router.delete("/admin/payments/{payment_id}")
async def delete_payment(
    payment_id: str,
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Delete a payment and recalculate customer dues."""
    payment = await db.payments.find_one({"id": payment_id, "tenant_id": tenant["id"]})
    if not payment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Payment not found")
    
    # Recalculate customer dues if payment was completed
    if payment.get("customer_id") and payment.get("payment_status") == PaymentStatus.COMPLETED:
        customer = await db.customers.find_one({"id": payment["customer_id"]})
        if customer:
            # Add back the payment amount to dues
            new_dues = customer.get("current_dues", 0) + payment["amount"]
            await db.customers.update_one(
                {"id": payment["customer_id"]},
                {"$set": {"current_dues": new_dues}}
            )
    
    await db.payments.delete_one({"id": payment_id})
    
    # Rebuild allocations if payment was COMPLETED
    customer_id = payment.get("customer_id")
    if customer_id and payment.get("payment_status") == PaymentStatus.COMPLETED:
        await _rebuild_customer_invoices_from_payments(db, tenant["id"], customer_id)
    
    return {"message": "Payment deleted successfully"}

@api_router.get("/admin/customers/{customer_id}/payments")
async def get_customer_payments(customer_id: str, limit: int = 10, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    """Get recent payments for a specific customer."""
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tenant["id"]})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    payments = await db.payments.find(
        {"tenant_id": tenant["id"], "customer_id": customer_id},
        {"_id": 0}
    ).sort("payment_date", -1).limit(limit).to_list(limit)
    
    return [Payment(**p) for p in payments]

# Invoice Routes
@api_router.get("/admin/invoices", response_model=List[Invoice])
async def get_invoices(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    invoices = await db.invoices.find({"tenant_id": tenant["id"]}, {"_id": 0}).to_list(10000)
    return [Invoice(**inv) for inv in invoices]

@api_router.post("/admin/invoices/generate-monthly")
async def generate_monthly_invoices(month: int = Body(...), year: int = Body(...), current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    tenant_id = tenant["id"]
    
    customers = await db.customers.find({"tenant_id": tenant_id, "is_active": True}, {"_id": 0}).to_list(10000)
    
    if not customers:
        return {"message": "No active customers found"}
    
    start_date = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    total_days = (end_date - start_date).days
    
    invoices_created = 0
    
    for customer in customers:
        existing = await db.invoices.find_one({
            "tenant_id": tenant_id,
            "customer_id": customer["id"],
            "month": month,
            "year": year
        })
        
        if existing:
            continue
        
        attendance = await db.attendance.find({
            "tenant_id": tenant_id,
            "customer_id": customer["id"],
            "date": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
        }, {"_id": 0}).to_list(10000)
        
        present_days = len(set(a["date"] for a in attendance if a.get("breakfast") or a.get("lunch") or a.get("dinner")))
        
        pause_requests = await db.pause_requests.find({
            "tenant_id": tenant_id,
            "customer_id": customer["id"],
            "status": PauseRequestStatus.APPROVED
        }, {"_id": 0}).to_list(10000)
        
        pause_days = 0
        for pause in pause_requests:
            pause_start = max(datetime.fromisoformat(pause["start_date"]), start_date)
            pause_end = min(datetime.fromisoformat(pause["end_date"]), end_date)
            if pause_start < pause_end:
                pause_days += (pause_end - pause_start).days
        
        meal_plan = await db.meal_plans.find_one({"id": customer["meal_plan_id"]})
        if not meal_plan:
            continue
        
        if meal_plan["billing_type"] == BillingType.MONTHLY:
            meal_charges = meal_plan["rate"]
        else:
            meal_charges = present_days * meal_plan["rate"]
        
        total_amount = meal_charges
        
        payments = await db.payments.find({
            "tenant_id": tenant_id,
            "customer_id": customer["id"],
            "payment_status": PaymentStatus.COMPLETED,
            "payment_date": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
        }, {"_id": 0}).to_list(10000)
        
        paid_amount = sum(p["amount"] for p in payments)
        due_amount = max(0, total_amount - paid_amount)
        
        invoice_count = await db.invoices.count_documents({"tenant_id": tenant_id})
        invoice_number = f"TMM-{tenant_id[:8]}-{year}-{invoice_count + 1:04d}"
        
        invoice_doc = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "customer_id": customer["id"],
            "customer_name": customer["full_name"],
            "invoice_number": invoice_number,
            "month": month,
            "year": year,
            "total_days": total_days,
            "present_days": present_days,
            "pause_days": pause_days,
            "meal_charges": meal_charges,
            "adjustments": 0,
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "due_amount": due_amount,
            "status": InvoiceStatus.PENDING,
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.invoices.insert_one(invoice_doc)
        
        # Rebuild allocations for this customer
        await _rebuild_customer_invoices_from_payments(db, tenant_id, customer["id"])
        
        
        invoices_created += 1
    
    return {"message": f"Generated {invoices_created} invoices for {month}/{year}"}

@api_router.post("/admin/invoices/generate-customer")
async def generate_customer_invoice(customer_id: str = Body(...), month: int = Body(...), year: int = Body(...), current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    """Generate invoice for a single customer."""
    tenant_info = tenant
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tenant_info["id"]})
    if not customer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    
    existing = await db.invoices.find_one({
        "tenant_id": tenant_info["id"],
        "customer_id": customer_id,
        "month": month,
        "year": year
    })
    
    if existing:
        return Invoice(**existing)
    
    start_date = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    total_days = (end_date - start_date).days
    
    attendance = await db.attendance.find({
        "tenant_id": tenant_info["id"],
        "customer_id": customer_id,
        "date": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    present_days = len(set(a["date"] for a in attendance if a.get("breakfast") or a.get("lunch") or a.get("dinner")))
    
    pause_requests = await db.pause_requests.find({
        "tenant_id": tenant_info["id"],
        "customer_id": customer_id,
        "status": PauseRequestStatus.APPROVED
    }, {"_id": 0}).to_list(10000)
    
    pause_days = 0
    for pause in pause_requests:
        pause_start = max(datetime.fromisoformat(pause["start_date"]), start_date)
        pause_end = min(datetime.fromisoformat(pause["end_date"]), end_date)
        if pause_start < pause_end:
            pause_days += (pause_end - pause_start).days
    
    meal_plan = await db.meal_plans.find_one({"id": customer["meal_plan_id"]})
    if not meal_plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Meal plan not found")
    
    if meal_plan["billing_type"] == BillingType.MONTHLY:
        meal_charges = meal_plan["rate"]
    else:
        meal_charges = present_days * meal_plan["rate"]
    
    total_amount = meal_charges
    
    payments = await db.payments.find({
        "tenant_id": tenant_info["id"],
        "customer_id": customer_id,
        "payment_status": PaymentStatus.COMPLETED,
        "payment_date": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    paid_amount = sum(p["amount"] for p in payments)
    due_amount = max(0, total_amount - paid_amount)
    
    invoice_count = await db.invoices.count_documents({"tenant_id": tenant_info["id"]})
    invoice_number = f"TMM-{tenant_info['id'][:8]}-{year}-{invoice_count + 1:04d}"
    
    invoice_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_info["id"],
        "customer_id": customer_id,
        "customer_name": customer["full_name"],
        "invoice_number": invoice_number,
        "month": month,
        "year": year,
        "total_days": total_days,
        "present_days": present_days,
        "pause_days": pause_days,
        "meal_charges": meal_charges,
        "adjustments": 0,
        "total_amount": total_amount,
        "paid_amount": paid_amount,
        "due_amount": due_amount,
            "status": InvoiceStatus.PENDING,
            "generated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.invoices.insert_one(invoice_doc)
    return Invoice(**invoice_doc)

@api_router.get("/admin/invoices/{invoice_id}/download")
async def download_invoice(invoice_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    """Download invoice as PDF. Only allowed for BASIC and above plans."""
    tenant_info = tenant
    
    plan = tenant_info.get("subscription_plan")
    if plan == SubscriptionPlan.FREE_TRIAL:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Invoice downloads are not available on Free Trial plan. Please upgrade to Basic or higher to download invoices."
        )
    
    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": tenant_info["id"]})
    if not invoice:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invoice not found")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    title = Paragraph(f"<b>Invoice {invoice['invoice_number']}</b>", styles['Heading1'])
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    details_data = [
        ["Invoice Number:", invoice['invoice_number']],
        ["Customer:", invoice['customer_name']],
        ["Month:", f"{invoice['month']}/{invoice['year']}"],
        ["Generated:", datetime.fromisoformat(invoice['generated_at']).strftime('%Y-%m-%d')],
    ]
    details_table = Table(details_data, colWidths=[2*inch, 4*inch])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    story.append(details_table)
    story.append(Spacer(1, 0.3*inch))
    
    charges_data = [
        ["Description", "Days", "Amount"],
        ["Meal Charges", str(invoice['present_days']), f"₹{invoice['meal_charges']:.2f}"],
        ["Total", "", f"₹{invoice['total_amount']:.2f}"],
        ["Paid", "", f"₹{invoice['paid_amount']:.2f}"],
        ["Due", "", f"₹{invoice['due_amount']:.2f}"],
    ]
    charges_table = Table(charges_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    charges_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -2), (-1, -1), colors.lightyellow),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(charges_table)
    story.append(Spacer(1, 0.3*inch))
    
    footer = Paragraph("<i>This is an electronically generated invoice. No signature is required.</i>", styles['Normal'])
    story.append(footer)
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{invoice['invoice_number']}.pdf"}
    )

# Staff Management Routes
@api_router.post("/admin/staff")
async def create_staff(staff_data: dict = Body(...), current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    staff_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    staff_doc = {
        "id": staff_id,
        "tenant_id": tenant["id"],
        "full_name": staff_data["full_name"],
        "gender": staff_data["gender"],
        "mobile": staff_data["mobile"],
        "email": staff_data.get("email"),
        "address": staff_data["address"],
        "role": staff_data["role"],
        "joining_date": staff_data["joining_date"],
        "salary": staff_data["salary"],
        "is_active": True,
        "total_advances": 0,
        "total_paid": 0,
        "created_at": now.isoformat()
    }
    
    await db.staff.insert_one(staff_doc)
    return staff_doc

@api_router.get("/admin/staff")
async def get_staff(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    staff_list = await db.staff.find({"tenant_id": tenant["id"]}, {"_id": 0}).to_list(1000)
    return staff_list

@api_router.get("/admin/staff/{staff_id}")
async def get_staff_detail(staff_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    staff = await db.staff.find_one({"id": staff_id, "tenant_id": tenant["id"]}, {"_id": 0})
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    payments = await db.staff_payments.find({"staff_id": staff_id, "tenant_id": tenant["id"]}, {"_id": 0}).to_list(1000)
    staff["payment_history"] = payments
    
    return staff

@api_router.put("/admin/staff/{staff_id}")
async def update_staff(staff_id: str, updates: dict = Body(...), current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    staff = await db.staff.find_one({"id": staff_id, "tenant_id": tenant["id"]})
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    if updates:
        await db.staff.update_one({"id": staff_id}, {"$set": updates})
    
    updated_staff = await db.staff.find_one({"id": staff_id}, {"_id": 0})
    return updated_staff

@api_router.delete("/admin/staff/{staff_id}")
async def delete_staff(staff_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    result = await db.staff.delete_one({"id": staff_id, "tenant_id": tenant["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Staff not found")
    return {"message": "Staff deleted successfully"}

@api_router.post("/admin/staff/{staff_id}/payment")
async def add_staff_payment(staff_id: str, payment_data: dict = Body(...), current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    staff = await db.staff.find_one({"id": staff_id, "tenant_id": tenant["id"]})
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    payment_doc = {
        "id": payment_id,
        "tenant_id": tenant["id"],
        "staff_id": staff_id,
        "staff_name": staff["full_name"],
        "amount": payment_data["amount"],
        "payment_type": payment_data["payment_type"],
        "notes": payment_data.get("notes"),
        "payment_date": now.isoformat(),
        "created_at": now.isoformat()
    }
    
    await db.staff_payments.insert_one(payment_doc)
    
    if payment_data["payment_type"] == "ADVANCE":
        await db.staff.update_one({"id": staff_id}, {"$inc": {"total_advances": payment_data["amount"]}})
    elif payment_data["payment_type"] == "SALARY":
        await db.staff.update_one({"id": staff_id}, {"$inc": {"total_paid": payment_data["amount"]}})
    
    return payment_doc

@api_router.get("/admin/payments/{payment_id}/invoice")
async def download_payment_invoice(
    payment_id: str,
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Download payment receipt/invoice as PDF."""
    payment = await db.payments.find_one({"id": payment_id, "tenant_id": tenant["id"]})
    if not payment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Payment not found")
    
    # Get customer details
    customer = None
    if payment.get("customer_id"):
        customer = await db.customers.find_one({"id": payment["customer_id"]}, {"_id": 0})
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a56db'),
        spaceAfter=30,
        alignment=1  # Center
    )
    title = Paragraph(f"<b>Payment Receipt</b>", title_style)
    story.append(title)
    story.append(Spacer(1, 0.2*inch))
    
    # Mess details
    mess_data = [
        [Paragraph(f"<b>{tenant['mess_name']}</b>", styles['Normal'])],
        [tenant.get('address', '')],
        [f"Mobile: {tenant.get('mobile', '')}"]
    ]
    mess_table = Table(mess_data, colWidths=[6*inch])
    mess_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(mess_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Receipt details
    receipt_data = [
        ["Receipt No:", payment['id'][:8].upper()],
        ["Date:", datetime.fromisoformat(payment['payment_date']).strftime('%d %B %Y, %I:%M %p')],
        ["Payment Method:", payment['payment_method']],
    ]
    
    if payment.get('transaction_id') or payment.get('razorpay_payment_id'):
        receipt_data.append([
            "Transaction ID:", 
            payment.get('razorpay_payment_id') or payment.get('transaction_id', '')
        ])
    
    if customer:
        receipt_data.extend([
            ["Customer Name:", customer['full_name']],
            ["Mobile:", customer['mobile']],
        ])
    
    receipt_table = Table(receipt_data, colWidths=[2*inch, 4*inch])
    receipt_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    story.append(receipt_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Amount section
    amount_data = [
        [Paragraph("<b>Description</b>", styles['Normal']), Paragraph("<b>Amount</b>", styles['Normal'])],
        ["Payment Received", f"₹{payment['amount']:,.2f}"],
    ]
    
    if payment.get('notes'):
        amount_data.append(["Notes:", payment['notes']])
    
    amount_data.append([
        Paragraph("<b>Total Amount</b>", styles['Normal']), 
        Paragraph(f"<b>₹{payment['amount']:,.2f}</b>", styles['Normal'])
    ])
    
    amount_table = Table(amount_data, colWidths=[4*inch, 2*inch])
    amount_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dbeafe')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(amount_table)
    story.append(Spacer(1, 0.5*inch))
    
    # Payment status badge
    status_text = payment.get('payment_status', 'COMPLETED')
    status_color = colors.HexColor('#10b981') if status_text == 'COMPLETED' else colors.HexColor('#f59e0b')
    status_para = Paragraph(
        f"<b>Status: {status_text}</b>",
        ParagraphStyle('Status', parent=styles['Normal'], textColor=status_color, fontSize=12)
    )
    story.append(status_para)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer = Paragraph(
        "<i>This is a computer-generated receipt and does not require a signature.<br/>"
        "Thank you for your payment!</i>",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=1)
    )
    story.append(footer)
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=receipt_{payment['id'][:8]}.pdf"}
    )


@api_router.get("/admin/payments/export")
async def export_payments(
    searchQuery: Optional[str] = None,
    paymentMethod: Optional[str] = None,
    paymentStatus: Optional[str] = None,
    dateFrom: Optional[str] = None,
    dateTo: Optional[str] = None,
    minAmount: Optional[float] = None,
    maxAmount: Optional[float] = None,
    current_user: dict = Depends(require_admin),
    tenant: dict = Depends(get_tenant_info)
):
    """Export payments to Excel with applied filters."""
    # Check plan permissions
    plan_config = PLAN_CONFIG.get(tenant.get("subscription_plan"))
    if not plan_config or not plan_config.get("allow_downloads", False):
        raise HTTPException(
            status_code=403,
            detail="Exports not allowed on your current plan. Please upgrade to Basic or higher."
        )
    
    # Build query with same filters as get_payments
    query = {"tenant_id": tenant["id"]}
    
    if searchQuery:
        search_lower = searchQuery.lower()
        customers = await db.customers.find(
            {"tenant_id": tenant["id"]},
            {"_id": 0, "id": 1, "full_name": 1}
        ).to_list(10000)
        
        matching_customer_ids = [
            c["id"] for c in customers 
            if search_lower in c["full_name"].lower()
        ]
        
        query["$or"] = [
            {"customer_id": {"$in": matching_customer_ids}},
            {"transaction_id": {"$regex": searchQuery, "$options": "i"}},
            {"razorpay_payment_id": {"$regex": searchQuery, "$options": "i"}},
            {"notes": {"$regex": searchQuery, "$options": "i"}}
        ]
    
    if paymentMethod and paymentMethod != "ALL":
        query["payment_method"] = paymentMethod
    
    if paymentStatus and paymentStatus != "ALL":
        query["payment_status"] = paymentStatus
    
    if dateFrom:
        date_from_obj = datetime.fromisoformat(dateFrom).replace(hour=0, minute=0, second=0, microsecond=0)
        query["payment_date"] = {"$gte": date_from_obj.isoformat()}
    
    if dateTo:
        date_to_obj = datetime.fromisoformat(dateTo).replace(hour=23, minute=59, second=59, microsecond=999999)
        if "payment_date" in query:
            query["payment_date"]["$lte"] = date_to_obj.isoformat()
        else:
            query["payment_date"] = {"$lte": date_to_obj.isoformat()}
    
    if minAmount is not None:
        query["amount"] = {"$gte": minAmount}
    
    if maxAmount is not None:
        if "amount" in query:
            query["amount"]["$lte"] = maxAmount
        else:
            query["amount"] = {"$lte": maxAmount}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Header styling
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Date", "Customer", "Amount", "Method", "Status", "Transaction ID", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row, payment in enumerate(payments, start=2):
        ws.cell(row=row, column=1, value=datetime.fromisoformat(payment['payment_date']).strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=payment.get('customer_name', 'N/A'))
        ws.cell(row=row, column=3, value=payment['amount'])
        ws.cell(row=row, column=4, value=payment['payment_method'])
        ws.cell(row=row, column=5, value=payment.get('payment_status', 'COMPLETED'))
        ws.cell(row=row, column=6, value=payment.get('razorpay_payment_id') or payment.get('transaction_id', ''))
        ws.cell(row=row, column=7, value=payment.get('notes', ''))
    
    # Adjust column widths
    column_widths = [20, 25, 15, 18, 12, 30, 40]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Add total row
    total_row = len(payments) + 2
    ws.cell(row=total_row, column=2, value="Total:").font = Font(bold=True)
    total_amount = sum(p['amount'] for p in payments if p.get('payment_status') == 'COMPLETED')
    ws.cell(row=total_row, column=3, value=total_amount).font = Font(bold=True)
    
    # Save to bytes - FIXED: Create new BytesIO and save properly
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Return StreamingResponse with proper headers
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=payments_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

# Reports Routes
@api_router.get("/admin/reports/customers-with-dues")
async def get_customers_with_dues(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customers = await db.customers.find({
        "tenant_id": tenant["id"],
        "is_active": True,
        "current_dues": {"$gt": 0}
    }, {"_id": 0}).to_list(1000)
    return customers

@api_router.get("/admin/reports/aging")
async def get_aging_report(current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    invoices = await db.invoices.find({
        "tenant_id": tenant["id"],
        "due_amount": {"$gt": 0}
    }, {"_id": 0}).to_list(1000)
    
    now = datetime.now(timezone.utc)
    aging_data = {"0-30": [], "31-60": [], "61-90": [], "90+": []}
    
    for invoice in invoices:
        days_old = (now - datetime.fromisoformat(invoice["generated_at"])).days
        customer = await db.customers.find_one({"id": invoice["customer_id"]}, {"_id": 0})
        
        invoice_data = {
            "customer_name": invoice["customer_name"],
            "customer_mobile": customer.get("mobile", "") if customer else "",
            "invoice_number": invoice["invoice_number"],
            "due_amount": invoice["due_amount"],
            "days_old": days_old
        }
        
        if days_old <= 30:
            aging_data["0-30"].append(invoice_data)
        elif days_old <= 60:
            aging_data["31-60"].append(invoice_data)
        elif days_old <= 90:
            aging_data["61-90"].append(invoice_data)
        else:
            aging_data["90+"].append(invoice_data)
    
    return aging_data

@api_router.get("/admin/reports/meal-consumption")
async def get_meal_consumption(month: int, year: int, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    start_date = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    attendance_records = await db.attendance.find({
        "tenant_id": tenant["id"],
        "date": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    consumption = defaultdict(lambda: {"breakfast": 0, "lunch": 0, "dinner": 0, "total": 0})
    
    for record in attendance_records:
        customer_id = record["customer_id"]
        if record.get("breakfast"):
            consumption[customer_id]["breakfast"] += 1
        if record.get("lunch"):
            consumption[customer_id]["lunch"] += 1
        if record.get("dinner"):
            consumption[customer_id]["dinner"] += 1
        consumption[customer_id]["total"] = consumption[customer_id]["breakfast"] + consumption[customer_id]["lunch"] + consumption[customer_id]["dinner"]
    
    result = []
    for customer_id, meals in consumption.items():
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if customer:
            result.append({
                "customer_name": customer["full_name"],
                "customer_mobile": customer["mobile"],
                **meals
            })
    
    return result

@api_router.get("/admin/reports/revenue")
async def get_revenue_report(months: int = 6, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    now = datetime.now(timezone.utc)
    revenue_data = []
    
    for i in range(months):
        target_date = now - timedelta(days=30*i)
        month_start = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if target_date.month == 12:
            month_end = target_date.replace(year=target_date.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            month_end = target_date.replace(month=target_date.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
        
        payments = await db.payments.find({
            "tenant_id": tenant["id"],
            "payment_status": PaymentStatus.COMPLETED,
            "payment_date": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()}
        }, {"_id": 0}).to_list(10000)
        
        total = sum(p["amount"] for p in payments)
        revenue_data.append({
            "month": month_start.strftime("%b %Y"),
            "revenue": total,
            "payment_count": len(payments)
        })
    
    return list(reversed(revenue_data))

# Export Routes
@api_router.get("/admin/export/customers")
async def export_customers(format: str = "csv", current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    plan_config = PLAN_CONFIG[tenant["subscription_plan"]]
    if not plan_config.get("allow_downloads", False):
        raise HTTPException(
            status_code=403, 
            detail="Downloads not allowed on your current plan. Please upgrade to Basic or higher to export data."
        )
    
    customers = await db.customers.find({"tenant_id": tenant["id"]}, {"_id": 0}).to_list(10000)
    
    if format == "csv":
        import csv
        from io import StringIO
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=["full_name", "mobile", "email", "meal_plan_name", "monthly_rate", "current_dues", "is_active"])
        writer.writeheader()
        for c in customers:
            writer.writerow({
                "full_name": c["full_name"],
                "mobile": c["mobile"],
                "email": c.get("email", ""),
                "meal_plan_name": c.get("meal_plan_name", ""),
                "monthly_rate": c["monthly_rate"],
                "current_dues": c["current_dues"],
                "is_active": c["is_active"]
            })
        return {"data": output.getvalue(), "filename": f"customers_{datetime.now().strftime('%Y%m%d')}.csv"}
    
    return {"message": "PDF export not implemented yet"}

# Subscription Management Routes
@api_router.post("/super-admin/subscriptions/renew/{tenant_id}")
async def renew_subscription(tenant_id: str, months: int = Body(...), current_user: dict = Depends(require_super_admin)):
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    current_end = datetime.fromisoformat(tenant["subscription_end"])
    new_end = current_end + timedelta(days=30*months)
    
    await db.tenants.update_one({"id": tenant_id}, {"$set": {
        "subscription_end": new_end.isoformat(),
        "status": TenantStatus.ACTIVE
    }})
    
    log_id = str(uuid.uuid4())
    await db.subscription_logs.insert_one({
        "id": log_id,
        "tenant_id": tenant_id,
        "previous_plan": tenant["subscription_plan"],
        "new_plan": tenant["subscription_plan"],
        "amount": PLAN_CONFIG[tenant["subscription_plan"]]["price"] * months,
        "payment_status": PaymentStatus.COMPLETED,
        "notes": f"Renewed for {months} months",
        "changed_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Subscription renewed successfully", "new_end_date": new_end.isoformat()}

@api_router.get("/super-admin/subscriptions/logs")
async def get_subscription_logs(current_user: dict = Depends(require_super_admin)):
    logs = await db.subscription_logs.find({}, {"_id": 0}).sort("changed_at", -1).to_list(1000)
    return logs

@api_router.post("/admin/tenants/{tenant_id}/renew")
async def admin_tenant_renew(
    tenant_id: str,
    subscription_plan: SubscriptionPlan = Body(...),
    months: int = Body(1),
    razorpay_order_id: str = Body(...),
    razorpay_payment_id: str = Body(...),
    current_user: dict = Depends(require_admin)
):
    """Finalize tenant subscription renewal after successful Razorpay payment."""
    if current_user.get("role") != UserRole.SUPER_ADMIN and current_user.get("tenant_id") != tenant_id:
        raise HTTPException(status_code=403, detail="Not authorized to renew this tenant")

    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    plan_cfg = PLAN_CONFIG.get(subscription_plan)
    if not plan_cfg:
        raise HTTPException(status_code=400, detail="Invalid subscription plan")

    amount_expected = plan_cfg.get("price", 0) * max(1, months)

    if amount_expected == 0:
        raise HTTPException(status_code=400, detail="Selected plan has zero price. Use manual renewal instead.")

    try:
        payment = razorpay_client.payment.fetch(razorpay_payment_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Razorpay verification failed: {str(e)}")

    status_str = payment.get("status")
    order_id = payment.get("order_id")
    paid_amount = int(payment.get("amount", 0)) / 100.0

    if status_str not in ("captured", "authorized"):
        raise HTTPException(status_code=400, detail=f"Payment not captured: status={status_str}")

    if order_id != razorpay_order_id:
        raise HTTPException(status_code=400, detail="Payment does not belong to provided order")

    if abs(paid_amount - float(amount_expected)) > 1.0:
        raise HTTPException(status_code=400, detail=f"Paid amount {paid_amount} does not match expected {amount_expected}")

    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    payment_doc = {
        "id": payment_id,
        "tenant_id": tenant_id,
        "customer_id": None,
        "customer_name": "Subscription Payment",
        "amount": paid_amount,
        "payment_method": PaymentMethod.RAZORPAY,
        "payment_status": PaymentStatus.COMPLETED,
        "transaction_id": payment.get("id"),
        "razorpay_order_id": razorpay_order_id,
        "razorpay_payment_id": razorpay_payment_id,
        "notes": f"Subscription renew: {subscription_plan} for {months} month(s)",
        "payment_date": now.isoformat(),
        "created_at": now.isoformat()
    }

    await db.payments.insert_one(payment_doc)

    now_dt = datetime.now(timezone.utc)
    new_end = now_dt + timedelta(days=30 * max(1, months))
    update_fields = {
        "subscription_plan": subscription_plan,
        "subscription_start": now_dt.isoformat(),
        "subscription_end": new_end.isoformat(),
        "status": TenantStatus.ACTIVE,
        "capacity": plan_cfg.get("capacity", tenant.get("capacity"))
    }

    await db.tenants.update_one({"id": tenant_id}, {"$set": update_fields})

    log_id = str(uuid.uuid4())
    await db.subscription_logs.insert_one({
        "id": log_id,
        "tenant_id": tenant_id,
        "previous_plan": tenant.get("subscription_plan"),
        "new_plan": subscription_plan,
        "amount": paid_amount,
        "payment_status": PaymentStatus.COMPLETED,
        "notes": f"Renewed via Razorpay payment {razorpay_payment_id}",
        "changed_at": now.isoformat()
    })

    return {"message": "Subscription renewed successfully", "new_end_date": new_end.isoformat(), "payment_id": payment_id}

# Customer Detail Route
@api_router.get("/admin/customers/{customer_id}/detail")
async def get_customer_detail(customer_id: str, current_user: dict = Depends(require_admin), tenant: dict = Depends(get_tenant_info)):
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tenant["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    payments = await db.payments.find({"customer_id": customer_id, "tenant_id": tenant["id"]}, {"_id": 0}).sort("payment_date", -1).to_list(100)
    
    invoices = await db.invoices.find({"customer_id": customer_id, "tenant_id": tenant["id"]}, {"_id": 0}).sort("generated_at", -1).to_list(100)
    
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    attendance = await db.attendance.find({
        "customer_id": customer_id,
        "tenant_id": tenant["id"],
        "date": {"$gte": thirty_days_ago.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    return {
        "customer": customer,
        "payments": payments,
        "invoices": invoices,
        "attendance_records": attendance
    }

# Razorpay webhook
@api_router.post("/webhooks/razorpay")
async def razorpay_webhook(request_body: dict = Body(...), x_razorpay_signature: Optional[str] = None):
    """Verify Razorpay webhook signature and process payment events."""
    secret = os.environ.get('RAZORPAY_WEBHOOK_SECRET')
    if not secret:
        raise HTTPException(status_code=500, detail="Webhook secret not configured")

    try:
        import json as _json
        body_str = _json.dumps(request_body, separators=(',', ':'), ensure_ascii=False)
        razorpay_client.utility.verify_webhook_signature(body_str, x_razorpay_signature, secret)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid webhook signature: {str(e)}")

    event = request_body.get('event')
    payload = request_body.get('payload', {})

    if event in ('payment.captured', 'payment.authorized'):
        payment_obj = payload.get('payment', {}).get('entity', {})
        payment_id = payment_obj.get('id')
        order_id = payment_obj.get('order_id')
        amount = int(payment_obj.get('amount', 0)) / 100.0

        existing = await db.payments.find_one({
            "$or": [
                {"razorpay_order_id": order_id},
                {"razorpay_payment_id": payment_id}
            ]
        })

        now = datetime.now(timezone.utc)
        if existing:
            await db.payments.update_one({"id": existing['id']}, {"$set": {
                "payment_status": PaymentStatus.COMPLETED,
                "razorpay_payment_id": payment_id,
                "payment_date": now.isoformat()
            }})
        else:
            payment_id_uuid = str(uuid.uuid4())
            payment_doc = {
                "id": payment_id_uuid,
                "tenant_id": None,
                "customer_id": None,
                "customer_name": "Subscription (webhook)",
                "amount": amount,
                "payment_method": PaymentMethod.RAZORPAY,
                "payment_status": PaymentStatus.COMPLETED,
                "transaction_id": payment_id,
                "razorpay_order_id": order_id,
                "razorpay_payment_id": payment_id,
                "notes": "Captured via webhook",
                "payment_date": now.isoformat(),
                "created_at": now.isoformat()
            }
            await db.payments.insert_one(payment_doc)

    return {"status": "ok"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    await initialize_default_data()
    logger.info("TrackMyMess API started")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()